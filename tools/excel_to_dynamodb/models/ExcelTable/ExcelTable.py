from typing import Optional, Union, Any, Dict, List, Tuple, Callable, Self, Generator
import openpyxl
import re

from pprint import pprint



class DataField:
    name:str = None
    value_org: Any = None
    value:Any = None
    attr:Dict = {}
    error:Exception = None

    def __init__(self, name:str, value:Any, attr:Dict[str, Any] = {}):
        self.name = name
        self.value_org = value
        self.value = value
        self.attr = attr
        self.error = None

    def serialize(self):
        return {
            "name" : self.name,
            "value_org" : self.value_org,
            "value" : self.value,
            **self.attr,
        }


    def __str__(self):
        return str(self.serialize())


class DataRecord(List[DataField]):

    def __init__(self, names:List[str], values:List[Any], attrs:List[Dict[str,Any]], attr_by_name:Dict[str, Dict[str, Any]]):
        super().__init__()

        for i in range(0, len(names)):
            # name = empty は除去
            if not names[i]:
                continue

            # fieldを生成してrecordを形成
            name:str = names[i]
            value:Any = values[i]
            attr = attrs[i] if i < len(attrs) else {}
            attr = { **attr, **attr_by_name.get(name, {}) }
            self.append(DataField(name=name, value=value, attr=attr))

    def map(self, func:callable) -> None:
        for field in self:
            try:
                field.error = None
                field.value = func(**field.serialize())
            except Exception as e:
                field.error = e


    def serialize(self) -> List[Dict]:
        return [
            x.serialize() for x in self
        ]

    def __str__(self):
        return str(self.serialize())

    def to_dict(self) -> Dict[str, List[Any] | Any]:
        # ※入れ物を作りながら値もセットすると、作り方によるが
        #   元の値が list の時にデータが正しくセットされなくなる。

        # 先に入れ物を作る。
        results:Dict[str, Any] = {}
        for x in self:
            if x.name in results.keys():
                results[x.name] = []
            else:
                results[x.name] = None

        # 値を入れる
        for x in self:
            if isinstance(results[x.name], list):
                results[x.name].append(x.value)
            else:
                results[x.name] = x.value
        
        # 返却
        return results




class DataTable():
    sheetname:str = ""
    tablename:str = ""
    records: List[DataRecord] = []

    def __init__(self, wb:openpyxl.workbook, sheetname:str, tablename:str, table_config:Dict):
        self.sheetname:str = sheetname
        self.tablename:str = tablename
        self.records: List[DataRecord] = []
        self._load2(wb=wb, sheetname=sheetname, table_config=table_config)
        # self._load(wb=wb, sheetname=sheetname, table_config=table_config)


    # def _load(self, wb:openpyxl.Workbook, sheetname:str, table_config:Dict) -> None:
    #     # 設定取得
    #     direction  = table_config["direction"]
    #     name_range = table_config["name_range"]
    #     data_range = table_config["data_range"]
    #     attr_range = table_config.get("attr_range", {})
    #     data_attr  = table_config.get("data_attr", {})

    #     # worksheet を読み込み
    #     worksheet = wb[sheetname]

    #     name_iter:Callable = None
    #     attr_iter:Callable = None
    #     data_iter:Callable = None
    #     if direction.lower() == 'horizonal':
    #         name_iter = getattr(worksheet, "iter_cols")
    #         data_iter = getattr(worksheet, "iter_cols")
    #         attr_iter = getattr(worksheet, "iter_rows")
    #     elif direction.lower() == 'vertical':
    #         name_iter = getattr(worksheet, "iter_rows")
    #         data_iter = getattr(worksheet, "iter_rows")
    #         attr_iter = getattr(worksheet, "iter_cols")
    #     else:
    #         raise ValueError(f"Unknown direction `{direction}` specified. `horizonal` or `vertical` is preferred.")

    #     # name_range を読み込み
    #     names:List[str|None] = [
    #         list(x[1:])
    #         for x in name_iter(**name_range, values_only=True,)
    #     ][0]

    #     # data_range を読み込み
    #     data_lines:List[Any] = [
    #         list(x[1:])
    #         for x in data_iter(**data_range, values_only=True,)
    #     ]
    #     # 空データをスキップ（除去）
    #     data_lines = [ x for x in data_lines if any(x)]

    #     # attr_range を読み込み
    #     attrs:List[Dict[str, Any]] = []
    #     if len(attr_range) > 0:
    #         attr_src:List[List[Any]] = [
    #             list(x)
    #             for x in attr_iter(**attr_range, values_only=True,)
    #         ]
    #         # attr を dict化
    #         attr_keys: List[str] = attr_src.pop(0)
    #         attrs:List[Dict[str, Any]] = [
    #             dict(zip(attr_keys, x)) for x in attr_src
    #         ]

    #     # 統合してデータ生成
    #     for data_line in data_lines:
    #         self.records.append(
    #             DataRecord(names=names, values=data_line, attrs=attrs, attr_by_name=data_attr)
    #         )


    def _load2(self, wb:openpyxl.Workbook, sheetname:str, table_config:Dict) -> None:
        # 設定取得
        data_ranges = table_config.get("data_ranges", {})
        data_attr   = table_config.get("data_attr", {})

        # worksheet オブジェクトを取得
        worksheet = wb[sheetname]

        # データ範囲設定ごとにデータを読み込み
        columner_lines: Dict[str, List[Any]] = {}
        for data_range in data_ranges:

            # 読み込み方向を取得
            direction: str = data_range["direction"].lower()
            if direction not in ["horizonal", "vertical"]:
                raise ValueError(f"Unknown direction `{direction}` specified. `horizonal` or `vertical` is preferred.")
            range_iter: Generator = {
                "horizonal" : worksheet.iter_cols,
                "vertical"  : worksheet.iter_rows,
            }[direction]

            # 読み込み範囲を取得
            range_def: Dict = {
                k: data_range.get(k, 0)
                for k in ["min_row", "min_col", "max_row", "max_col"]
            }

            # 指定範囲を読み込み
            data_lines:List[Any] = [
                list(x)
                for x in range_iter(**range_def, values_only=True,)
            ]

            # header を取得
            headers: List[str]= data_range.get("header", None)
            if not headers:
                # 設定にない場合は読み込みデータの先頭要素を header と見なしてから取得
                headers = [ x.pop(0) for x in data_lines ]

            # header のうち name, value は小文字化
            headers = [ (x.lower() if x.lower() in ["name", "value"] else x) for x in headers]

            # header が不足かつ末尾が value の場合は増幅して補完
            if len(headers) < len(data_range) and headers[-1] == "value":
                headers.extend(
                    ["value"] * (len(data_range) - len(headers) - 2 )
                )

            # データを列指向テーブルに変換
            for data_line in data_lines:
                if len(headers) > 0:
                    header = headers.pop(0)

                # value の場合はリスト化（複数件対応のため）
                if header.lower() == "value":
                    if "value" not in columner_lines.keys():
                        columner_lines[header] = []
                    columner_lines[header].append(data_line)
                else:
                    # value 以外は追加代入
                    columner_lines[header] = data_line


        # name, value の必須チェック
        if "name" not in columner_lines.keys():
            raise KeyError("Column name `name` does not exist.")

        if "value" not in columner_lines.keys():
            raise KeyError("Column name `value` does not exist.")

        # name と attr を抜き出し
        names:List[str] = columner_lines["name"]

        attr_keys: List[str] = [ k for k in columner_lines.keys() if k not in ("name", "value") ]
        attrs:List[Dict[str, Any]] = []
        for i in range(0, len(names)):
            attrs.append({
                k:columner_lines[k][i] for k in attr_keys
            })

        # value の空行はスキップ（除去）
        columner_lines["value"] = [ x for x in columner_lines["value"] if any(x)]

        # 統合してレコード生成
        for data_line in columner_lines["value"]:
            self.records.append(
                DataRecord(names=names, values=data_line, attrs=attrs, attr_by_name=data_attr)
            )

        # 終了
        return


    def map(self, func:callable) -> None:
        for record in self.records:
            record.map(func)


    def serialize(self) -> Dict[str, Any]:
        return {
            "sheetname": self.sheetname,
            "tablename": self.tablename,
            "records"  : [ x.serialize() for x in self.records ],
        }

    def __str__(self) -> str:
        return str(self.serialize())

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheetname": self.sheetname,
            "tablename": self.tablename,
            "records"  : [ x.to_dict() for x in self.records ],
        }



class DataTables(List[DataTable]):

    def __init__(self, file:str, load_configs:List[Dict]):
        super().__init__()
        self._load(file=file, load_configs=load_configs)


    def _load(self, file:str, load_configs:List[Dict]) -> None:

        # worksheet を読み込み
        wb = openpyxl.load_workbook(
            filename=file,
            read_only=False,    # True にすると `iter_cols` で AttributeError になる
            data_only=True
        )

        for load_config in load_configs:
            # シート名を取得
            sheetname_pattern:str = load_config["sheetname"]
            sheetname_regex = sheetname_pattern.replace("*", r"(.*)")
            sheetnames:List[str] = [
                x for x in wb.sheetnames
                if re.fullmatch(sheetname_regex, x)
            ]

            # シート＋テーブル定義ごとに読み込み
            for sheetname in sheetnames:
                for tablename, table_config in load_config["tables"].items():
                    table_key:str = f"{sheetname}/{tablename}"
                    self.append(DataTable(
                        wb=wb,
                        sheetname=sheetname,
                        tablename=tablename,
                        table_config=table_config
                    ))

        # 正常終了
        return

    def filter(self, sheetname:str=None, tablename:str=None) -> List[DataTable]:
        return [
            x for x in self
            if ( sheetname in [x.sheetname, None])
            and ( tablename in [x.tablename, None])
        ]

    def tablenames(self) -> List[str]:
        """tablename 一覧（重複なし）"""
        return list(set([ table.tablename for table in self ]))

    def sheetname(self) -> List[str]:
        """sheetname 一覧（重複なし）"""
        return list(set([ table.sheetname for table in self ]))

    def map(self, func:callable) -> None:
        for table in self:
            table.map(func=func)


    def serialize(self) -> List[Dict[str, Any]]:
        return [
            x.serialize()
            for x in self
        ]

    def __str__(self) -> str:
        return str(self.serialize())


    def to_dict(self) -> List[Dict[str, Any]]:
        return [
            x.to_dict()
            for x in self
        ]

