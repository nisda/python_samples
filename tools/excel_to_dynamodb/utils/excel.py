from typing import Dict, List
import openpyxl


def batch_load(file:str, load_configs:Dict[str, Dict]) -> Dict[str, List[Dict[str, str|None]]]:
    def __load(file:str, load_config:Dict) -> List[Dict[str, str|None]]:
        # 設定取得
        sheetname = load_config["sheetname"]
        direction = load_config["direction"]
        range = load_config["range"]

        # worksheet からデータ取得
        wb = openpyxl.load_workbook(
            filename=file,
            read_only=False,    # True にすると `iter_cols` で AttributeError
            data_only=True
        )
        worksheet = wb[sheetname]

        worksheet_iter = None
        if direction.lower() == 'horizonal':
            worksheet_iter = getattr(worksheet, "iter_cols")
        elif direction.lower() == 'vertical':
            worksheet_iter = getattr(worksheet, "iter_rows")

        lines:List[Dict] = [
            list(x)
            for x in worksheet_iter(**range,values_only=True,)
        ]

        header:List[str] = lines.pop(0)
        table: List[Dict[str, str]] = [
            dict(zip(header, x)) for x in lines
        ]

        return table



    ret = {
        k: __load(file=file, load_config=v) for k,v in load_configs.items()
    }
    return ret



def batch_load2(file:str, load_configs:Dict[str, Dict]) -> Dict[str, List[Dict[str, str|None]]]:
    def __load(file:str, load_config:Dict) -> List[Dict[str, str|None]]:
        # 設定取得
        sheetname = load_config["sheetname"]
        direction = load_config["direction"]
        data_range = load_config["data"]
        attr_range = load_config.get("attr", None)

        # worksheetを読み込み
        wb = openpyxl.load_workbook(
            filename=file,
            read_only=False,    # True にすると `iter_cols` で AttributeError
            data_only=True
        )
        worksheet = wb[sheetname]

        # 読み込み方向の設定
        data_iter = None
        if direction.lower() == 'horizonal':
            data_iter = getattr(worksheet, "iter_cols")
            attr_iter = getattr(worksheet, "iter_rows")
        elif direction.lower() == 'vertical':
            data_iter = getattr(worksheet, "iter_rows")
            attr_iter = getattr(worksheet, "iter_cols")


        # 属性読み込み
        if attr_range is not None:
            attr_lines:List[Dict] = [
                list(x)
                for x in attr_iter(**attr_range, values_only=True,)
            ]

        # データ読み込み
        data_lines:List[Dict] = [
            list(x)
            for x in data_iter(**data_range, values_only=True,)
        ]

        # データをテーブル化
        data_header:List[str] = data_lines.pop(0)
        table: List[Dict[str, str]] = [
            dict(zip(data_header, x)) for x in data_lines
        ]

        # attr をテーブル化
        attr_header:List[str] = attr_lines.pop(0)
        attrs: List[Dict[str, str]] = [
            dict(zip(attr_header, x)) for x in attr_lines
        ]


        # dict of list に変換
        temp = []
        for rec in table:
            new_rec: List[Dict] = []
            for i in range(0, len(rec.keys())):
                name:str = list(rec.keys())[i]
                value:str|None = list(rec.values())[i]
                new_rec.append({
                    "name": name,
                    "value": value
                })
            temp.append(new_rec)
        records = temp

        return records



    ret = {
        k: __load(file=file, load_config=v) for k,v in load_configs.items()
    }
    return ret

