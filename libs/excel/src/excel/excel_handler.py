from typing import Union,Optional, List,Dict,Tuple, Any
import openpyxl
import io

class ExcelHandler:


    __file: str | bytes = None
    __wb: openpyxl.workbook.workbook.Workbook = None

    def __init__(self):
        pass

    def load(self, file:str | bytes=None, read_only:bool=True, data_only:bool=True):
        if isinstance(file, str):
            filename:str = file
        if isinstance(file, bytes):
            filename:io.BytesIO = io.BytesIO(file)
        self.__wb = openpyxl.load_workbook(filename=filename, read_only=read_only, data_only=data_only)

    def sheetnames(self) -> list[str]:
        return self.__wb.sheetnames

    def worksheets(self) -> list[openpyxl.worksheet.worksheet.Worksheet]:
        return self.__wb.worksheets

    def worksheet(self, sheet_name:str) -> openpyxl.worksheet.worksheet.Worksheet:
        if sheet_name not in self.sheetnames():
            return None
        return self.__wb[sheet_name]

    def row_values(self, sheet_name:str, min_row:int=1, min_col:int=1, max_row:int=0, max_col:int=0, first_list_as_header:bool=False):
        table:List[Dict] = []
        for line in self.worksheet(sheet_name=sheet_name).iter_rows(
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
            values_only=False,
        ):
            lines = []
            for cell in line:
                lines.append(cell.value)
            table.append(lines)

        return table




