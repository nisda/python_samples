from typing import Union,Optional, List,Dict,Tuple, Any, overload, Iterator, Self
import os
import io
import fnmatch
from enum import StrEnum, auto
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from .excel_worksheet import ExcelWorksheet


class ExcelWorkbook:

    __path:Path = None
    __wb: Workbook = None


    @property
    def path(self) -> Path:
        return self.__path

    @property
    def sheetnames(self) -> List[str]:
        """ワークシート名の一覧"""
        return self.__wb.sheetnames


    def __init__(self, file:str|bytes|None=None, read_only:bool=True, data_only:bool=True):
        self.__wb = None
        self.__file = None

        if file is not None:
            self.load(file=file, read_only=read_only, data_only=data_only)


    def load(self, file:str|bytes|io.BytesIO, read_only:bool=True, data_only:bool=True) -> Self:
        self.__path =  None

        if isinstance(file, str):
            filename:str = file
            self.__path = Path(file)
        elif isinstance(file, io.BytesIO):
            filename:io.BytesIO = file
        elif isinstance(file, bytes):
            filename:io.BytesIO = io.BytesIO(file)

        self.__wb = openpyxl.load_workbook(filename=filename, read_only=read_only, data_only=data_only)
        return self


    def new(self, sheetname="Sheet") -> Self:
        # ロード済みチェック
        if self.__wb is not None:
            raise Exception("The workbook has already been loaded.")
        # 新規Book作成
        self.__wb = openpyxl.Workbook()
        self.__wb.active.title = sheetname
        return self



    def save(self):
        # 保存（上書き専用）
        if self.__wb is None:
            raise Exception("Workbook is not loaded.")
        if self.__path is None:
            raise Exception("FilePath is not set.")
        self.__wb.save(filename=self.__path)


    def save_as(self, path:str, overwrite:bool=False) -> Self:

        # ロード済みチェック
        if self.__wb is None:
            raise Exception("Workbook is not loaded.")


        if self.__path is None:
            # 初回保存は可。自分自身を返却
            self.__wb.save(filename=path)
            self.__path = Path(path)
            return self

        elif self.__path.resolve() == Path(path).resolve():
            # パスが同一の場合は保存可。自分自身を返却
            self.__wb.save(filename=path)
            return self

        elif not os.path.isfile(path):
            # ファイル非存在時は保存して新規インスタンスを返却
            self.__wb.save(filename=path)
            return ExcelWorkbook(file=path, read_only=False, data_only=False)

        elif os.path.isfile(path) and overwrite:
            # パスが異なる＆ファイル存在時は上書き可否チェック
            self.__wb.save(filename=path)
            return ExcelWorkbook(file=path, read_only=False, data_only=False)

        else:
            # ファイル存在かつ上書き不可
            raise FileExistsError("File `{path}` already exists.")



    def worksheets(self, pattern:str="*") -> List[ExcelWorksheet]:
        """ワークシート（複数）取得"""
        return [
            ExcelWorksheet(self.__wb[sheetname])
            for sheetname in self.sheetnames
            if fnmatch.fnmatch(name=sheetname, pat=pattern)
        ]


    def worksheet(self, sheetname:str) -> ExcelWorksheet:
        """ワークシート取得"""
        if sheetname in self.sheetnames:
            return ExcelWorksheet(worksheet=self.__wb[sheetname])
        else:
            return None


    def create_sheet(self, sheetname:str, index:int=None) -> ExcelWorksheet:
        """シート追加"""

        # 重複チェック
        if sheetname in self.sheetnames:
            raise KeyError(f"Worksheet `{sheetname}` is already exists.")

        sheet = self.__wb.create_sheet(title=sheetname, index=index)
        return ExcelWorksheet(worksheet=sheet)


    def remove_sheet(self, sheetname:str):
        """シート削除"""
        if sheetname not in self.sheetnames:
            raise KeyError(f"Worksheet `{sheetname}` does not exist.")

        self.__wb.remove(self.__wb[sheetname])


    def copy_sheet(self, src:str, dest:str, index:int=None) -> ExcelWorksheet:
        """シート複製"""

        # コピー元の存在チェック
        if src not in self.sheetnames:
            raise KeyError(f"Worksheet `{src}` is not found.")

        # コピー先の重複チェック
        if dest in self.sheetnames:
            raise KeyError(f"Worksheet `{dest}` is already exists.")

        # コピー実行
        sheet = self.__wb.copy_worksheet(self.__wb[src])
        sheet.title = dest

        # 位置調整
        if isinstance(index, int):
            offset:int = index - self.__wb.index(sheet)
            self.__wb.move_sheet(sheet, offset=offset)

        return ExcelWorksheet(worksheet=sheet)


    def sort_sheet(self, name_order:List[str]) -> List[str]:
        """
        シートの並べ替え
        ----
        name_order の通りに並べ替える。
        name_order に含まれないシートは後ろに寄せる。
        """

        # 指定されたnameの後ろから先頭に移動していく
        for sheetname in reversed(name_order):
            if sheetname not in self.sheetnames:
                continue
            ws:Worksheet = self.__wb[sheetname]
            self.__wb.move_sheet(ws, offset=-self.__wb.index(ws))

        return self.sheetnames





    # @classmethod
    # def __convert_coordinate(cls, coord:int|str|None, default):
    #     """Excelの座標表現を数値化"""
    #     if isinstance(coord, str):
    #         return openpyxl.utils.column_index_from_string(coord)
    #     return int(coord or default)


    # @classmethod
    # def __tuples_to_dict(cls, tuples:List[Tuple[Any, Any]]) -> Dict:
    #     """Key-Value がセットになった tuple リストを dict 化"""

    #     def __convert_scalar_or_list(value):
    #         """scalar/list調整: listかつ要素が1つであればscalarにする"""
    #         if isinstance(value, list) and len(value) == 1:
    #             return value[0]
    #         else:
    #             return value

    #     temp = defaultdict(list)
    #     for key, value in tuples:
    #         temp[key].append(value)
    #     return {k: __convert_scalar_or_list(v) for k,v in temp.items()}


    # def get_values(
    #         self,
    #         sheet:Union[Worksheet, str], header:bool=False, skip_blank:bool=True,
    #         min_col:int|str="A", min_row:int=1, max_col:int|str=None,max_row:int=None,
    #         read_direction:Union[Direction, str]=Direction.Vertical,
    #         return_oriented:Union[Oriented, str]=Oriented.Row,
    #     ) -> Union[List, Dict]:
    #     """指定シートからデータを読み込み"""

    #     #-----------------------------
    #     # パラメータ調整
    #     #-----------------------------

    #     # ワークシート取得
    #     ws:Worksheet = self.__wb[sheet] if isinstance(sheet, str) else sheet

    #     # データ形式を統一
    #     min_col = self.__convert_coordinate(coord=min_col, default=1)
    #     min_row = self.__convert_coordinate(coord=min_row, default=1)
    #     max_col = self.__convert_coordinate(coord=max_col, default=0)
    #     max_row = self.__convert_coordinate(coord=max_row, default=0)
    #     read_direction = Direction(read_direction)
    #     return_oriented = Oriented(return_oriented)


    #     #-----------------------------
    #     # データ取得
    #     #-----------------------------

    #     # データ読み込み
    #     # ※ReadOnlySheet は iter_cols が利用できないため、iter_rows で読み取る。
    #     lines:List[str] = [
    #         [ value for value in line ]
    #         for line in ws.iter_rows(
    #             min_row=min_row,
    #             min_col=min_col,
    #             max_row=max_row,
    #             max_col=max_col,
    #             values_only=True,
    #         )
    #     ]

    #     if read_direction == Direction.Horizontal:
    #         # 横方向指定の場合は行/列を転置
    #         lines = list(map(list, zip(*lines)))

    #     # 空行を除去
    #     if skip_blank:
    #         lines = [ x for x in lines if any(x) ]


    #     #-----------------------------
    #     # ヘッダ処理 ＆ 返却形式（列指向/行指向）の調整
    #     #-----------------------------
    #     if header:
    #         # ヘッダあり

    #         columns: List[str] = lines.pop(0)
    #         lines = [
    #             self.__tuples_to_dict(tuples=zip(columns, x))
    #             for x in lines
    #         ]

    #         if return_oriented == Oriented.Column:
    #             lines = {key: [d[key] for d in lines] for key in lines[0].keys()}

    #     else:
    #         # ヘッダなし
    #         if return_oriented == Oriented.Column:
    #             lines = list(map(list, zip(*lines)))


    #     return lines


    # def put_values(            
    #         self,
    #         data:List|Dict,
    #         sheet:Union[Worksheet, str],
    #         min_col:int|str="A", min_row:int=1, max_col:int|str=None,max_row:int=None,
    #         read_direction:Union[Direction, str]=Direction.Vertical,
    #     ):
    #     """指定シートにデータを書き込み"""


    #     #-----------------------------
    #     # パラメータ調整
    #     #-----------------------------

    #     # ワークシート取得
    #     ws:Worksheet = self.__wb[sheet] if isinstance(sheet, str) else sheet

    #     # データ形式を統一
    #     min_col = self.__convert_coordinate(coord=min_col, default=1)
    #     min_row = self.__convert_coordinate(coord=min_row, default=1)
    #     max_col = self.__convert_coordinate(coord=max_col, default=0)
    #     max_row = self.__convert_coordinate(coord=max_row, default=0)
    #     read_direction = Direction(read_direction)


    #     #-----------------------------
    #     # データを整形
    #     #-----------------------------
    #     if len(data) == 0:
    #         # 空データだったらそのまま終了
    #         return

    #     records: List = []

    #     # if isinstance(data, dict):
    #     #     # 列指向データ
    #     #     keys = list(data.keys())
    #     #     records = [
    #     #         [ x for k in keys
    #     #         for d in data
    #     #     ]
    #     #     records.insert(0, keys) # 先頭にカラム行をINSERT

    #     # elif isinstance(data, (list, tuple)):
    #     #     if len()
    #     # else:
    #     #     raise TypeError(f"Type=`{type(data)}` is not supprted.")




