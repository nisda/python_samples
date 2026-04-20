from typing import Union,Optional, List,Dict,Tuple, Any, overload, Iterator, Self
import os
import io
import fnmatch
from enum import StrEnum, auto
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


class __StrEnumEx(StrEnum):
    @classmethod
    def _missing_(cls, value):
        if isinstance(value, str):
            for member in cls:
                if member.value.casefold() == value.casefold():
                    return member
        return super()._missing_(value)

class Direction(__StrEnumEx):
    Vertical = auto()
    Horizontal = auto()


class Oriented(__StrEnumEx):
    Row = auto()
    Column = auto()


class ExcelWorksheet:

    __worksheet: Worksheet = None


    @property
    def sheetname(self) -> str:
        return self.__worksheet.title

    @sheetname.setter
    def sheetname(self, value):
        self.__worksheet.title = value


    def __init__(self, worksheet:Worksheet):
        self.__worksheet = worksheet


    @classmethod
    def __convert_coordinate(cls, coord:int|str|None, default):
        """Excelの座標表現を数値化"""
        if isinstance(coord, str):
            return openpyxl.utils.column_index_from_string(coord)
        return int(coord or default)


    @classmethod
    def __tuples_to_dict(cls, tuples:List[Tuple[Any, Any]]) -> Dict:
        """Key-Value がセットになった tuple リストを dict 化"""

        def __convert_scalar_or_list(value):
            """scalar/list調整: listかつ要素が1つであればscalarにする"""
            if isinstance(value, list) and len(value) == 1:
                return value[0]
            else:
                return value

        temp = defaultdict(list)
        for key, value in tuples:
            temp[key].append(value)
        return {k: __convert_scalar_or_list(v) for k,v in temp.items()}


    def get_values(
            self, header:bool=False, skip_blank:bool=True,
            min_col:int|str="A", min_row:int=1, max_col:int|str=None,max_row:int=None,
            read_direction:Union[Direction, str]=Direction.Vertical,
            return_oriented:Union[Oriented, str]=Oriented.Row,
        ) -> Union[List, Dict]:
        """指定シートからデータを読み込み"""

        #-----------------------------
        # パラメータ調整
        #-----------------------------

        # ワークシート取得
        ws:Worksheet = self.__worksheet

        # データ形式を統一
        min_col = self.__convert_coordinate(coord=min_col, default=1)
        min_row = self.__convert_coordinate(coord=min_row, default=1)
        max_col = self.__convert_coordinate(coord=max_col, default=0)
        max_row = self.__convert_coordinate(coord=max_row, default=0)
        read_direction = Direction(read_direction)
        return_oriented = Oriented(return_oriented)


        #-----------------------------
        # データ取得
        #-----------------------------

        # データ読み込み
        # ※ReadOnlySheet は iter_cols が利用できないため、iter_rows で読み取る。
        lines:List[str] = [
            [ value for value in line ]
            for line in ws.iter_rows(
                min_row=min_row,
                min_col=min_col,
                max_row=max_row,
                max_col=max_col,
                values_only=True,
            )
        ]

        if read_direction == Direction.Horizontal:
            # 横方向指定の場合は行/列を転置
            lines = list(map(list, zip(*lines)))

        # 空行を除去
        if skip_blank:
            lines = [ x for x in lines if any(x) ]


        #-----------------------------
        # ヘッダ処理 ＆ 返却形式（列指向/行指向）の調整
        #-----------------------------
        if header:
            # ヘッダあり

            columns: List[str] = lines.pop(0)
            lines = [
                self.__tuples_to_dict(tuples=zip(columns, x))
                for x in lines
            ]

            if return_oriented == Oriented.Column:
                lines = {key: [d[key] for d in lines] for key in lines[0].keys()}

        else:
            # ヘッダなし
            if return_oriented == Oriented.Column:
                lines = list(map(list, zip(*lines)))


        return lines


    def put_values(            
            self,
            data:List|Dict,
            min_col:int|str="A", min_row:int=1,
            max_col:int|str=None,max_row:int=None,
            direction:Union[Direction, str]=Direction.Vertical,
        ) -> None:
        """指定シートにデータを書き込み"""


        #-----------------------------
        # パラメータ調整
        #-----------------------------

        # ワークシート取得
        ws:Worksheet = self.__worksheet

        # データ形式を統一
        min_col = self.__convert_coordinate(coord=min_col, default=1)
        min_row = self.__convert_coordinate(coord=min_row, default=1)
        max_col = self.__convert_coordinate(coord=max_col, default=0)
        max_row = self.__convert_coordinate(coord=max_row, default=0)
        direction = Direction(direction)


        #-----------------------------
        # データを整形
        #-----------------------------
        records:List[List[Any]] = []
        if len(data) > 0:
            if isinstance(data, dict):
                # 列指向、列名あり
                records.append(list(data.keys()))
                records.extend(list(map(list, zip(*data.values()))))
            elif isinstance(data[0], list):
                # 行指向、列名なし
                records = data
            elif isinstance(data[0], dict):
                # 行指向、列名あり
                keys:List[Any] = list(data[0].keys())
                records = [
                    [ line.get(k, None) for k in keys ]
                    for line in data
                ]
                records.insert(0, keys)


        if direction == Direction.Horizontal:
            # 行列転置
            records = list(map(list, zip(*records)))

        #-----------------------------
        # 書き込み
        #-----------------------------
        records = self.__resize_records(
            records = records,
            width   = max_col - min_col + 1,
            height  = max_row - min_row + 1,
        )

        for r_idx, row in enumerate(records):
            for c_idx, value in enumerate(row):
                # 座標を指定して書き込み
                ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)

        #-----------------------------
        # 書き込み
        #-----------------------------
        return None


    @classmethod
    def __resize_records(cls, records:List[List[Any]], width:int=0, height:int=0, default:Any=None):

        if height <= 0 :
            height = len(records)
        if width <= 0:
            width = len(records[0]) if len(records) > 0 else 0

        result = []
        for row in range(0, height):
            base:List[Any] = [default] * width
            if row < len(records):
                rec = records[row][0:len(base)]
                base[0:len(rec)] = rec
                result.append(base)
            else:
                result.append(base)

        return result
