from typing import Union,Optional, List,Dict,Tuple, Any
import openpyxl
import sys
from datetime import datetime
import shutil

from  openpyxl.worksheet.worksheet import Worksheet as Worksheet
INPUT_FILE_PATH  = "../file/sample_lock_in.xlsx"
OUTPUT_FILE_PATH = "../file/sample_lock_out.xlsx"

if __name__ == '__main__':
    shutil.copy(src=INPUT_FILE_PATH, dst=OUTPUT_FILE_PATH)


    wb = openpyxl.load_workbook(filename=OUTPUT_FILE_PATH, read_only=False, data_only=False)

    #--------------------------------------
    #   シート名一覧
    #--------------------------------------
    print()
    print("-----------------------")
    print("シート名一覧")
    print("-----------------------")
    # シート名一覧
    ws_names = wb.sheetnames
    print(ws_names)

    #--------------------------------------
    #   データ取得
    #--------------------------------------
    print()
    print("-----------------------")
    print("データ取得")
    print("-----------------------")

    # データ取得（非保護セル）
    ws:Worksheet = wb["保護お試し"]
    value = ws.cell(2, 2).value
    print(value)

    # データ取得（保護セル）
    ws = wb["保護お試し"]
    value = ws.cell(15, 2).value
    print(value)

    # データ取得（非表示シート）
    ws = wb["非表示シート"]
    value = ws.cell(1, 1).value
    print(value)


    #--------------------------------------
    #   データ書き込み
    #--------------------------------------
    print()
    print("-----------------------")
    print("データ書き込み")
    print("-----------------------")

    # データ書き込み（非保護セル）
    ws:Worksheet = wb["保護お試し"]
    cell = ws.cell(4, 4)
    cell.value = f"書き込み={datetime.now().isoformat()}"

    # データ書き込み（非保護セル） => エラーにならない！？
    ws:Worksheet = wb["保護お試し"]
    cell = ws.cell(16, 3)
    cell.value = f"書き込み={datetime.now().isoformat()}"

    # データ書き込み（非表示シート）
    ws = wb["非表示シート"]
    cell = ws.cell(3, 3)
    cell.value = f"書き込み={datetime.now().isoformat()}"



    #--------------------------------------
    #   シートのリネーム
    #--------------------------------------
    print()
    print("-----------------------")
    print("シートのリネーム")
    print("-----------------------")

    ws = wb["非表示シート"]
    ws.title = "非表示シートNew"

    wb.save(filename=OUTPUT_FILE_PATH)

